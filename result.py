from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
maxRow = ws.max_row
for row in range(2,maxRow+1):
    hours = ws['B'+str(row)].value
    rate = ws['C'+str(row)].value
    if isinstance(hours, int) and isinstance(rate, int): #type() vietā
        salary = hours * rate
        if salary > 3000:
            total += 1
print("People with salary >3000€:",total)
wb.close()
